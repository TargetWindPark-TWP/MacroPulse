"""
MacroPulse — Excel 報告生成器
每次執行產出一份時間戳記命名的 .xlsx，包含：
  Sheet 1 「總覽」        — 所有指標最新值 + YoY + MoM
  Sheet 2 「通膨」        — 通膨類歷史數據
  Sheet 3 「經濟成長」    — 成長類歷史數據
  Sheet 4 「勞動市場」    — 就業類歷史數據
  Sheet 5 「利率與債券」  — 利率類歷史數據
  Sheet 6 「信心指數」    — 情緒類歷史數據
  Sheet 7 「信用與違約」  — 信用類歷史數據
  Sheet 8 「房市」        — 房市類歷史數據
"""
from __future__ import annotations
import json
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# ─── 色彩常數 ─────────────────────────────────────────────────────────────────
C = {
    "header_bg":   "1F3864",   # 深藍（標題列）
    "header_fg":   "FFFFFF",
    "cat_inflation":"FFE0E0",  # 通膨 - 淡紅
    "cat_growth":  "E0F5F4",   # 成長 - 淡青
    "cat_labor":   "FFFBDD",   # 就業 - 淡黃
    "cat_rates":   "FFE8E0",   # 利率 - 淡橘
    "cat_sentiment":"E8F8F0",  # 信心 - 淡綠
    "cat_credit":  "FFE0E0",   # 信用 - 淡紅
    "cat_housing": "EEE4FF",   # 房市 - 淡紫
    "positive":    "C6EFCE",   # 正值背景
    "positive_fg": "276221",
    "negative":    "FFC7CE",   # 負值背景
    "negative_fg": "9C0006",
    "neutral":     "FFEB9C",   # 中性背景
    "neutral_fg":  "9C6500",
    "subheader":   "D6E4F7",   # 副標題
    "alt_row":     "F5F9FF",   # 交替行
    "white":       "FFFFFF",
    "border":      "B0C4DE",
}

CAT_COLORS = {
    "inflation": C["cat_inflation"],
    "growth":    C["cat_growth"],
    "labor":     C["cat_labor"],
    "rates":     C["cat_rates"],
    "sentiment": C["cat_sentiment"],
    "credit":    C["cat_credit"],
    "housing":   C["cat_housing"],
}

CAT_LABELS = {
    "inflation": "🔥 通膨指標",
    "growth":    "📈 經濟成長",
    "labor":     "👷 勞動市場",
    "rates":     "🏦 利率與債券",
    "sentiment": "💭 信心指數",
    "credit":    "⚠️ 信用與違約",
    "housing":   "🏠 房市",
}

CATEGORY_ORDER = ["inflation", "growth", "labor", "rates", "sentiment", "credit", "housing"]

SHEET_NAMES = {
    "inflation": "通膨",
    "growth":    "經濟成長",
    "labor":     "勞動市場",
    "rates":     "利率與債券",
    "sentiment": "信心指數",
    "credit":    "信用與違約",
    "housing":   "房市",
}

# 利率類指標（YoY/MoM 單位為 pp 百分點，而非 %）
RATE_SERIES = {
    "FEDFUNDS", "DGS2", "DGS10", "DGS30", "T10Y2Y",
    "MORTGAGE30US", "DRCCLACBS", "DRSFRMACBS", "DRCONGACBS",
    "BAMLH0A0HYM2", "BAMLC0A0CM", "UNRATE", "VIXCLS",
    "NFCI", "STLFSI4",
}


# ─── 樣式工廠 ─────────────────────────────────────────────────────────────────

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def make_font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def make_border(color=C["border"]) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def make_align(horizontal="center", vertical="center", wrap=False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical,
                     wrap_text=wrap)


# ─── 自動欄寬 ─────────────────────────────────────────────────────────────────

def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ─── 數值格式化 ───────────────────────────────────────────────────────────────

def fmt_raw(value, series_id: str, raw_unit: str) -> str:
    if value is None:
        return "N/A"
    v = float(value)
    if "Thousands" in raw_unit or "thousands" in raw_unit:
        return f"{v:,.0f}"
    if "Millions" in raw_unit or "millions" in raw_unit:
        return f"{v:,.0f}"
    if "Bil." in raw_unit:
        return f"{v:,.1f}"
    return f"{v:.4g}"


def fmt_pct(value) -> str:
    if value is None:
        return "N/A"
    return f"{float(value):.2f}%"


def fmt_pp(value) -> str:
    """百分點差值（利率類）"""
    if value is None:
        return "N/A"
    v = float(value)
    return f"{v:+.3f} pp"


def yoy_label(series_id: str) -> str:
    return "YoY (pp)" if series_id in RATE_SERIES else "YoY %"


def mom_label(series_id: str) -> str:
    return "MoM (pp)" if series_id in RATE_SERIES else "MoM %"


def yoy_fmt(value, series_id: str) -> str:
    if value is None:
        return "N/A"
    return fmt_pp(value) if series_id in RATE_SERIES else fmt_pct(value)


def mom_fmt(value, series_id: str) -> str:
    if value is None:
        return "N/A"
    return fmt_pp(value) if series_id in RATE_SERIES else fmt_pct(value)


# ─── 條件色彩（根據指標語義判斷好壞）──────────────────────────────────────────

def value_color(series_id: str, field: str, value) -> Optional[tuple]:
    """回傳 (bg_color, fg_color) 或 None"""
    if value is None:
        return None
    v = float(value)

    if field == "yoy":
        # 通膨指標 YoY：越高越紅
        if series_id in {"CPIAUCSL","CPILFESL","PCEPI","PCEPILFE","PPIACO","PPIFIS"}:
            if v > 3.5: return (C["negative"], C["negative_fg"])
            if v > 2.5: return (C["neutral"],  C["neutral_fg"])
            return (C["positive"], C["positive_fg"])
        # 殖利率倒掛：負值紅
        if series_id == "T10Y2Y":
            if v < 0:   return (C["negative"], C["negative_fg"])
            if v < 0.5: return (C["neutral"],  C["neutral_fg"])
            return (C["positive"], C["positive_fg"])
        # 失業率 YoY：上升為紅
        if series_id == "UNRATE":
            if v > 0.5: return (C["negative"], C["negative_fg"])
            if v > 0:   return (C["neutral"],  C["neutral_fg"])
            return (C["positive"], C["positive_fg"])

    if field == "raw":
        if series_id == "UNRATE":
            if v > 5: return (C["negative"], C["negative_fg"])
            if v > 4: return (C["neutral"],  C["neutral_fg"])
            return (C["positive"], C["positive_fg"])
        if series_id == "T10Y2Y":
            if v < 0: return (C["negative"], C["negative_fg"])
        if series_id in {"DRCCLACBS","DRSFRMACBS","BAMLH0A0HYM2","BAMLC0A0CM"}:
            if v > 3: return (C["negative"], C["negative_fg"])
        if series_id in {"NFCI","STLFSI4"}:
            if v > 0.5: return (C["negative"], C["negative_fg"])
            if v < -0.5: return (C["positive"], C["positive_fg"])

    return None


def apply_color(cell, series_id: str, field: str, value):
    colors = value_color(series_id, field, value)
    if colors:
        cell.fill = make_fill(colors[0])
        cell.font = make_font(bold=True, color=colors[1], size=10)


# ─── Sheet 1：總覽 ───────────────────────────────────────────────────────────

def build_summary_sheet(ws, all_data: dict, indicators: list, generated_at: str):
    ws.title = "📊 總覽"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # ── 標題 ──
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"MacroPulse 總體經濟指標報告　　生成時間：{generated_at}"
    title_cell.font      = make_font(bold=True, color=C["header_fg"], size=13)
    title_cell.fill      = make_fill(C["header_bg"])
    title_cell.alignment = make_align("left")
    ws.row_dimensions[1].height = 28

    # ── 表頭 ──
    headers = [
        "類別", "指標名稱", "Series ID", "頻率",
        "最新日期", "原始值", "原始單位",
        "YoY / 年增率", "YoY 單位",
        "MoM / 月增率", "MoM 單位",
        "數據來源",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = make_font(bold=True, color=C["header_fg"], size=10)
        cell.fill      = make_fill("2E5090")
        cell.alignment = make_align("center")
        cell.border    = make_border()
    ws.row_dimensions[2].height = 20

    # ── 數據列 ──
    row = 3
    for ind in indicators:
        sid  = ind["id"]
        data = all_data.get(sid)
        cat  = ind.get("category", "")
        bg   = CAT_COLORS.get(cat, C["white"])

        raw_val = data.get("latest_raw_value") if data else None
        yoy_val = data.get("latest_yoy_value") if data else None
        mom_val = data.get("latest_mom_value") if data else None
        raw_date = data.get("latest_raw_date", "") if data else ""

        yoy_str = yoy_fmt(yoy_val, sid)
        mom_str = mom_fmt(mom_val, sid)

        row_data = [
            CAT_LABELS.get(cat, cat),
            ind.get("name", sid),
            sid,
            ind.get("frequency", ""),
            raw_date,
            fmt_raw(raw_val, sid, ind.get("raw_unit", "")),
            ind.get("raw_unit", ""),
            yoy_str,
            yoy_label(sid),
            mom_str,
            mom_label(sid),
            ind.get("source", "FRED"),
        ]

        alt = (row % 2 == 0)
        row_bg = C["alt_row"] if alt else C["white"]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = make_fill(bg if col == 1 else row_bg)
            cell.alignment = make_align(
                "left" if col in (2, 7, 9, 11, 12) else "center"
            )
            cell.border = make_border()
            cell.font   = make_font(size=10)

            # 條件色彩
            if col == 6:
                apply_color(cell, sid, "raw", raw_val)
            elif col == 8:
                apply_color(cell, sid, "yoy", yoy_val)

        ws.row_dimensions[row].height = 18
        row += 1

    auto_width(ws, min_w=10, max_w=45)
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["L"].width = 28


# ─── Sheet 2-8：各類別歷史數據 ────────────────────────────────────────────────

def build_category_sheet(ws, category: str, all_data: dict, indicators: list,
                          generated_at: str, max_rows: int = 120):
    cat_indicators = [i for i in indicators if i.get("category") == category]
    if not cat_indicators:
        return

    ws.title = SHEET_NAMES.get(category, category)
    ws.sheet_view.showGridLines = False
    cat_bg = CAT_COLORS.get(category, C["white"])
    cat_label = CAT_LABELS.get(category, category)

    # ── 標題 ──
    total_cols = 1 + len(cat_indicators) * 3
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    t = ws["A1"]
    t.value      = f"{cat_label}　歷史數據　　生成時間：{generated_at}"
    t.font       = make_font(bold=True, color=C["header_fg"], size=12)
    t.fill       = make_fill(C["header_bg"])
    t.alignment  = make_align("left")
    ws.row_dimensions[1].height = 26

    # ── 指標名稱合併標題（第 2 列）──
    col = 2
    for ind in cat_indicators:
        ws.merge_cells(
            start_row=2, start_column=col,
            end_row=2,   end_column=col + 2
        )
        c = ws.cell(row=2, column=col, value=f"{ind['name']} ({ind['id']})")
        c.font      = make_font(bold=True, color=C["header_fg"], size=10)
        c.fill      = make_fill("2E5090")
        c.alignment = make_align("center")
        c.border    = make_border()
        col += 3
    ws.cell(row=2, column=1, value="日期").font = make_font(
        bold=True, color=C["header_fg"], size=10)
    ws.cell(row=2, column=1).fill      = make_fill("2E5090")
    ws.cell(row=2, column=1).alignment = make_align("center")
    ws.cell(row=2, column=1).border    = make_border()
    ws.row_dimensions[2].height = 20

    # ── 子欄位標題（第 3 列）──
    ws.cell(row=3, column=1, value="日期")
    ws.cell(row=3, column=1).font      = make_font(bold=True, size=9)
    ws.cell(row=3, column=1).fill      = make_fill(C["subheader"])
    ws.cell(row=3, column=1).alignment = make_align("center")
    ws.cell(row=3, column=1).border    = make_border()

    col = 2
    for ind in cat_indicators:
        sid = ind["id"]
        sub_headers = ["原始值", yoy_label(sid), mom_label(sid)]
        for sh in sub_headers:
            c = ws.cell(row=3, column=col, value=sh)
            c.font      = make_font(bold=True, size=9)
            c.fill      = make_fill(C["subheader"])
            c.alignment = make_align("center")
            c.border    = make_border()
            col += 1
    ws.row_dimensions[3].height = 18

    # ── 整合歷史數據（以日期為行）──
    # 收集所有日期
    date_set = set()
    series_data: dict[str, dict] = {}
    for ind in cat_indicators:
        sid  = ind["id"]
        data = all_data.get(sid, {})
        raw_obs = {o["date"]: o["value"] for o in (data.get("raw_observations") or [])}
        yoy_obs = {o["date"]: o["value"] for o in (data.get("yoy_observations") or [])}
        mom_obs = {o["date"]: o["value"] for o in (data.get("mom_observations") or [])}
        series_data[sid] = {"raw": raw_obs, "yoy": yoy_obs, "mom": mom_obs}
        date_set.update(raw_obs.keys())

    sorted_dates = sorted(date_set)[-max_rows:]  # 最近 max_rows 筆

    # ── 填入數據 ──
    ws.freeze_panes = "B4"

    for r_idx, dt in enumerate(sorted_dates):
        row = 4 + r_idx
        alt = (r_idx % 2 == 0)
        row_bg = C["alt_row"] if alt else C["white"]

        # 日期欄
        dc = ws.cell(row=row, column=1, value=dt)
        dc.font      = make_font(size=9)
        dc.fill      = make_fill(cat_bg)
        dc.alignment = make_align("center")
        dc.border    = make_border()

        col = 2
        for ind in cat_indicators:
            sid = ind["id"]
            sd  = series_data[sid]

            raw_v = sd["raw"].get(dt)
            yoy_v = sd["yoy"].get(dt)
            mom_v = sd["mom"].get(dt)

            for field, val in [("raw", raw_v), ("yoy", yoy_v), ("mom", mom_v)]:
                c = ws.cell(row=row, column=col)

                if field == "raw":
                    display = fmt_raw(val, sid, ind.get("raw_unit", ""))
                elif field == "yoy":
                    display = yoy_fmt(val, sid)
                else:
                    display = mom_fmt(val, sid)

                c.value     = display
                c.font      = make_font(size=9)
                c.fill      = make_fill(row_bg)
                c.alignment = make_align("center")
                c.border    = make_border()

                # 條件色彩
                if field in ("raw", "yoy") and val is not None:
                    apply_color(c, sid, field, val)

                col += 1

        ws.row_dimensions[row].height = 16

    auto_width(ws, min_w=9, max_w=18)
    ws.column_dimensions["A"].width = 14


# ─── 主函數 ───────────────────────────────────────────────────────────────────

def generate_excel(
    all_data: dict,
    indicators: list,
    output_dir: Path,
    max_history_rows: int = 120,
) -> Path:
    """
    生成 Excel 報告，回傳檔案路徑。
    檔名格式：MacroPulse_YYYY-MM-DD.xlsx
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    today   = date.today().isoformat()

    wb = Workbook()
    wb.remove(wb.active)   # 移除預設空白 sheet

    # Sheet 1：總覽
    ws_summary = wb.create_sheet("📊 總覽")
    build_summary_sheet(ws_summary, all_data, indicators, now_str)

    # Sheet 2-8：各類別
    for cat in CATEGORY_ORDER:
        ws_cat = wb.create_sheet(SHEET_NAMES.get(cat, cat))
        build_category_sheet(
            ws_cat, cat, all_data, indicators,
            now_str, max_rows=max_history_rows,
        )

    # 儲存
    filename = f"MacroPulse_{today}.xlsx"
    path = output_dir / filename
    wb.save(str(path))

    return path
